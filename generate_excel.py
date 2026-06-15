#!/usr/bin/env python3
# generate_excel.py - 주간 엑셀 리포트 (WoW: 직전주 마지막 거래일 → 이번주 마지막 거래일)
#   2026-06-15: 주간비교 로직을 '주 내 첫/마지막'에서 'WoW 마지막 거래일끼리'로 변경.
#               휴일은 스냅샷 데이터 유무로 자동 판별(get_wow_pair). 전주대비 %p, 제외종목 표시 추가.

import os, json
from datetime import datetime, timedelta
from pathlib import Path
from config.etfs import SHEETS

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print('openpyxl 미설치: pip install openpyxl')
    raise

from fetch_daily import load_snapshot, date_str, get_wow_pair

DATA_DIR   = Path('data/snapshots')
OUTPUT_DIR = Path('output')
OUTPUT_DIR.mkdir(exist_ok=True)

CLR = {
    'header_bg':  '1F3864', 'header_fg':  'FFFFFF',
    'sheet_bg':   '2E75B6', 'sheet_fg':   'FFFFFF',
    'etf_bg':     'D6E4F0',
    'new_in':     'C6EFCE', 'removed':    'FFCCCC', 'increased':  'FFEB9C',
    'row_odd':    'F8F9FA', 'row_even':   'FFFFFF', 'border':     'BDD7EE',
}

def make_fill(hex_color):  return PatternFill('solid', fgColor=hex_color)
def make_border(color='BDD7EE'):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)
def make_font(bold=False, color='000000', size=10):
    return Font(bold=bold, color=color, size=size, name='맑은 고딕')

def safe_sheet_title(name):
    for ch in '\\/*?:[]':
        name = name.replace(ch, '·')
    return name[:31]

# ── WoW 비교 ────────────────────────────────────────────────────────
def detect_changes_pair(prev_holdings, this_holdings):
    """직전주(prev) → 이번주(this) 보유 비교. (신규, 제외, 증가, this_map, prev_map, has_prev)"""
    prev_map = {h['name']: h for h in (prev_holdings or [])}
    this_map = {h['name']: h for h in (this_holdings or [])}
    has_prev = bool(prev_map)
    if not has_prev:
        return [], [], [], this_map, prev_map, False
    new_in    = [n for n in this_map if n not in prev_map]
    removed   = [n for n in prev_map if n not in this_map]
    increased = [n for n in this_map
                 if n in prev_map and this_map[n]['weight'] > prev_map[n]['weight'] + 0.3]
    return new_in, removed, increased, this_map, prev_map, True

def write_sheet_tab(wb, sheet_cfg, prev_snap, this_snap):
    sheet_name = sheet_cfg['name']
    ws = wb.create_sheet(title=safe_sheet_title(sheet_name))
    etfs = sheet_cfg['etfs']
    ws.cell(1, 1, f'[ {sheet_name} ]').font = Font(bold=True, size=13, color=CLR['header_fg'], name='맑은 고딕')
    ws.cell(1, 1).fill = make_fill(CLR['header_bg'])
    ws.cell(1, 1).alignment = Alignment(horizontal='center')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    headers = ['종목명', '비중(%)', '수량', '전주대비', '변화']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(2, ci, h)
        c.fill = make_fill(CLR['sheet_bg']); c.font = make_font(bold=True, color='FFFFFF')
        c.alignment = Alignment(horizontal='center')
    cur_row = 3
    for etf in etfs:
        etf_name = etf['name']
        is_passive = etf.get('is_passive', False)
        label = f"{'[패시브] ' if is_passive else ''}{etf_name}"
        c = ws.cell(cur_row, 1, label)
        c.fill = make_fill(CLR['etf_bg']); c.font = make_font(bold=True, size=10)
        c.alignment = Alignment(horizontal='left', indent=1)
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=5)
        cur_row += 1

        new_in, removed, increased, this_map, prev_map, has_prev = detect_changes_pair(
            prev_snap.get(etf_name, []), this_snap.get(etf_name, []))

        if not this_map:
            ws.cell(cur_row, 1, '데이터 없음').font = make_font(color='888888')
            cur_row += 1; cur_row += 1; continue
        if not has_prev:
            n = ws.cell(cur_row, 1, '📍 직전주 비교데이터 없음 — 현재 보유만 표시(WoW는 다음주부터)')
            n.font = make_font(color='888888'); cur_row += 1

        # 현재 보유 (비중 내림차순)
        holdings = sorted(this_map.values(), key=lambda x: x['weight'], reverse=True)
        for i, h in enumerate(holdings):
            name = h['name']
            if has_prev and name in new_in:
                row_fill = make_fill(CLR['new_in']); change_label = '🟢 신규'; delta_txt = '신규'
            elif has_prev and name in increased:
                row_fill = make_fill(CLR['increased']); change_label = '🔼 증가'
                delta_txt = f"{h['weight'] - prev_map[name]['weight']:+.2f}%p"
            else:
                row_fill = make_fill(CLR['row_odd'] if i % 2 == 0 else CLR['row_even'])
                change_label = '-'
                if has_prev and name in prev_map:
                    d = h['weight'] - prev_map[name]['weight']
                    delta_txt = f"{d:+.2f}%p" if abs(d) >= 0.005 else '–'
                else:
                    delta_txt = '–'
            _write_row(ws, cur_row, [name, h['weight'], h['qty'], delta_txt, change_label], row_fill)
            cur_row += 1

        # 이번주 빠진 종목(제외) — 직전주엔 있었으나 사라짐
        if has_prev and removed:
            for name in sorted(removed, key=lambda n: prev_map[n]['weight'], reverse=True):
                ph = prev_map[name]
                _write_row(ws, cur_row, [name, ph['weight'], ph['qty'], '제외', '🔴 제외'],
                           make_fill(CLR['removed']))
                cur_row += 1
        cur_row += 1

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.freeze_panes = 'A3'

def _write_row(ws, row, vals, fill):
    for ci, val in enumerate(vals, 1):
        c = ws.cell(row, ci, val)
        c.fill = fill; c.font = make_font(); c.border = make_border()
        if ci == 2:
            c.number_format = '0.00"%"'; c.alignment = Alignment(horizontal='center')
        elif ci == 3:
            c.number_format = '#,##0'; c.alignment = Alignment(horizontal='right')
        elif ci in (4, 5):
            c.alignment = Alignment(horizontal='center')
        else:
            c.alignment = Alignment(horizontal='left', indent=1)

def write_summary_tab(wb, prev_snap, this_snap, prev_ref, this_ref):
    ws = wb.create_sheet(title='📋 주간요약', index=0)
    ws.cell(1, 1, '📊 주간 ETF 신호 요약 (WoW)').font = Font(bold=True, size=14, name='맑은 고딕', color=CLR['header_fg'])
    ws.cell(1, 1).fill = make_fill(CLR['header_bg']); ws.cell(1, 1).alignment = Alignment(horizontal='center')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    prev_lbl = date_str(prev_ref) if prev_ref else '비교데이터 없음'
    ws.cell(2, 1, f'비교: {prev_lbl}  →  {date_str(this_ref)} (각 주 마지막 거래일)').font = make_font(color='666666')
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    headers = ['섹터', 'ETF명', '신규 편입', '비중 증가', '제외', '신호']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(3, ci, h)
        c.fill = make_fill(CLR['sheet_bg']); c.font = make_font(bold=True, color='FFFFFF')
        c.alignment = Alignment(horizontal='center')
    cur_row = 4
    for sheet_cfg in SHEETS:
        sheet_name = sheet_cfg['name']
        for etf in sheet_cfg['etfs']:
            etf_name = etf['name']
            new_in, removed, increased, _, _, has_prev = detect_changes_pair(
                prev_snap.get(etf_name, []), this_snap.get(etf_name, []))
            if not (new_in or removed or increased):
                continue
            if len(new_in) >= 2:   signal = '🔴 주목'
            elif new_in:           signal = '🟡 신규'
            elif increased:        signal = '🔼 증가'
            else:                  signal = '⬇️ 제외'
            row_data = [sheet_name, etf_name, ', '.join(new_in[:3]) or '-',
                        ', '.join(increased[:3]) or '-', ', '.join(removed[:3]) or '-', signal]
            for ci, val in enumerate(row_data, 1):
                c = ws.cell(cur_row, ci, val)
                c.font = make_font(); c.border = make_border()
                c.alignment = Alignment(horizontal='left' if ci <= 2 else 'center', wrap_text=True)
                if signal == '🔴 주목':
                    c.fill = make_fill(CLR['new_in'])
            cur_row += 1
    if cur_row == 4:
        ws.cell(4, 1, '이번주 변동 신호 없음').font = make_font(color='888888')
    ws.column_dimensions['A'].width = 14; ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 25; ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 20; ws.column_dimensions['F'].width = 10
    ws.freeze_panes = 'A4'

def main():
    from zoneinfo import ZoneInfo
    today = datetime.now(ZoneInfo('Asia/Seoul')).date()
    while today.weekday() >= 5:
        today -= timedelta(days=1)
    prev_ref, this_ref = get_wow_pair(today)
    print(f'=== 주간 엑셀(WoW): 직전주 {prev_ref} → 이번주 {this_ref} ===')
    if not this_ref:
        print('⚠️ 이번주 데이터 없음 — 엑셀 생성 생략')
        return None
    this_snap = load_snapshot(this_ref)
    prev_snap = load_snapshot(prev_ref) if prev_ref else {}
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    write_summary_tab(wb, prev_snap, this_snap, prev_ref, this_ref)
    for sheet_cfg in SHEETS:
        write_sheet_tab(wb, sheet_cfg, prev_snap, this_snap)
    filename = f'etf_weekly_{date_str(this_ref)}.xlsx'
    path = OUTPUT_DIR / filename
    wb.save(path)
    print(f'✅ 엑셀 저장 완료: {path}')
    return str(path)

if __name__ == '__main__':
    main()
